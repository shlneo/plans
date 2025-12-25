import io
import xml.etree.ElementTree as ET
from ..models import Plan, EconMeasure, EconExec
from sqlalchemy.orm import joinedload
from ..views import get_cumulative_econ_metrics
        
def export_xml_single(plan: Plan):
    """Экспорт одного плана в XML с тремя разделами и титульными данными."""

    def build_title_xml(plan):
        """Формирует XML-узел титульного листа."""
        title = ET.Element("title")

        approved = ET.SubElement(title, "approved_by")
        approved.text = (
            "СОГЛАСОВАНО: Департамент по энергоэффективности Госстандарта. "
        )

        ministrystr = plan.organization.ministry or "Министерство (концерн, государственный комитет)"
        confirmed = ET.SubElement(title, "confirmed_by")
        confirmed.text = (
            f"УТВЕРЖДАЮ {ministrystr}. "
        )

        ET.SubElement(title, "header").text = "ПЛАН МЕРОПРИЯТИЙ ПО ЭНЕРГОСБЕРЕЖЕНИЮ"
        ET.SubElement(title, "organization_name").text = str(plan.organization.name or "")
        ET.SubElement(title, "year_label").text = f"на {plan.year} год"

        targets = ET.SubElement(title, "targets")
        ET.SubElement(targets, "label").text = "Целевые показатели: показатель энергосбережения"
        ET.SubElement(targets, "details").text = (
            f"показатель энергосбережения - {plan.energy_saving}% "
            f"(задание по экономии ТЭР - {plan.share_fuel} т у.т.); "
            f"доля местных ТЭР в КПТ - {plan.saving_fuel}%; "
            f"доля местных ТЭР в КПТ - {plan.share_energy}%."
        )
        return title

    def build_part1_xml(plan):
        """
        Формирует XML-раздел 'Часть 1' с показателями использования ТЭР.
        """
        part1 = ET.Element("part1")
        
        ET.SubElement(part1, "title").text = "Часть 1. Показатели использования топливно-энергетических ресурсов"
        
        previous_group = None
        for usage in sorted(plan.indicators_usage, key=lambda u: (u.indicator.Group, u.indicator.RowN)):
            group_value = usage.indicator.Group if usage.indicator.Group != previous_group else ""
            previous_group = usage.indicator.Group

            row = ET.SubElement(part1, "row")
            ET.SubElement(row, "group").text = str(group_value or "")
            ET.SubElement(row, "name").text = str(usage.indicator.name or "-")
            ET.SubElement(row, "unit").text = str(getattr(usage.indicator.unit, "name", "") or "")
            ET.SubElement(row, "prev_year").text = str(usage.QYearPrev or 0)
            ET.SubElement(row, "curr_year").text = str(usage.QYearCurr or 0)
            ET.SubElement(row, "next_year").text = str(usage.QYearNext or 0)
            ET.SubElement(row, "change").text = str((usage.QYearNext or 0) - (usage.QYearCurr or 0))
        
        return part1

    def build_part2_xml(plan):
        """
        Формирует XML-раздел 'Часть 2' с мероприятиями по реализации основных направлений энергосбережения.
        """
        part2 = ET.Element("part2")
        
        ET.SubElement(part2, "title").text = f"Часть 2. Мероприятия по реализации основных направлений энергосбережения на {plan.year} год"
        
        for idx, measure in enumerate(sorted(plan.econ_measures, key=lambda u: u.direction.code), start=1):
            row = ET.SubElement(part2, "row")
            ET.SubElement(row, "number").text = str(idx)
            ET.SubElement(row, "code").text = str(measure.direction.code or "")
            ET.SubElement(row, "name").text = str(measure.direction.name or "")
            ET.SubElement(row, "year_econ").text = str(float(measure.year_econ or 0))
            ET.SubElement(row, "estim_econ").text = str(float(measure.estim_econ or 0))
        
        return part2

    def build_part3_xml(plan, get_cumulative_econ_metrics, EconExec):
        """
        Формирует XML-раздел 'Часть 3' с мероприятиями по увеличению использования местных ТЭР.
        """
        part3 = ET.Element("part3")
        ET.SubElement(part3, "title").text = "Часть 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов"

        local_execs = [e for e in plan.econ_execes if e.is_local]
        non_local_execs = [e for e in plan.econ_execes if not e.is_local]

        def add_section(title, execs, start_number=1):
            section = ET.SubElement(part3, "section", {"title": title})
            for idx, econ in enumerate(execs, start=start_number):
                row = ET.SubElement(section, "row")
                ET.SubElement(row, "number").text = str(idx)
                ET.SubElement(row, "code").text = str(econ.econ_measures.direction.code if econ.econ_measures and econ.econ_measures.direction else "")
                ET.SubElement(row, "name").text = str(econ.name if hasattr(econ, "name") else "")
                ET.SubElement(row, "unit").text = str(econ.econ_measures.direction.unit.name if econ.econ_measures and econ.econ_measures.direction and econ.econ_measures.direction.unit else "")
                ET.SubElement(row, "volume").text = str(getattr(econ, "Volume", 0))
                ET.SubElement(row, "eff_tut").text = str(getattr(econ, "EffTut", 0))
                ET.SubElement(row, "eff_rub").text = str(getattr(econ, "EffRub", 0))
                ET.SubElement(row, "expected_quarter").text = str(getattr(econ, "ExpectedQuarter", ""))
                ET.SubElement(row, "eff_curr_year").text = str(getattr(econ, "EffCurrYear", 0))
                ET.SubElement(row, "payback").text = str(getattr(econ, "Payback", 0))
                ET.SubElement(row, "volume_fin").text = str(getattr(econ, "VolumeFin", 0))
                ET.SubElement(row, "budget_state").text = str(getattr(econ, "BudgetState", 0))
                ET.SubElement(row, "budget_rep").text = str(getattr(econ, "BudgetRep", 0))
                ET.SubElement(row, "budget_loc").text = str(getattr(econ, "BudgetLoc", 0))
                ET.SubElement(row, "budget_other").text = str(getattr(econ, "BudgetOther", 0))
                ET.SubElement(row, "money_own").text = str(getattr(econ, "MoneyOwn", 0))
                ET.SubElement(row, "money_loan").text = str(getattr(econ, "MoneyLoan", 0))
                ET.SubElement(row, "money_other").text = str(getattr(econ, "MoneyOther", 0))
            return start_number + len(execs)

        next_number = add_section("Раздел 2.1 Мероприятия по экономии ТЭР (первоначальная ред.)", non_local_execs, 1)
        add_section("Раздел 3.1. Мероприятия по увеличению использования местных ТЭР (первоначальная ред.)", local_execs, next_number)

        quarters = [
            ("Январь–Март", "jan_mar"),
            ("Январь–Июнь", "jan_jun"),
            ("Январь–Сентябрь", "jan_sep"),
            ("Январь–Декабрь", "jan_dec")
        ]
        local_totals = get_cumulative_econ_metrics(plan.id, True)
        non_local_totals = get_cumulative_econ_metrics(plan.id, False)

        totals_elem = ET.SubElement(part3, "totals")
        for q_label, q_key in quarters:
            quarter = ET.SubElement(totals_elem, "quarter", {"name": q_label})
            ET.SubElement(quarter, "eff_curr_year").text = str(local_totals[q_key]["eff_curr_year"] + non_local_totals[q_key]["eff_curr_year"])
            ET.SubElement(quarter, "volume_fin").text = str(local_totals[q_key]["volume_fin"] + non_local_totals[q_key]["volume_fin"])

        return part3

    root = ET.Element("plan", attrib={
        "id": str(plan.id),
        "year": str(plan.year or "")
    })

    root.append(build_title_xml(plan))
    root.append(build_part1_xml(plan))
    root.append(build_part2_xml(plan))
    root.append(build_part3_xml(plan, get_cumulative_econ_metrics, EconExec))

    def prettify(elem, level=0):
        indent = "  "
        i = "\n" + level * indent
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + indent
            for child in elem:
                prettify(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = i
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i

    prettify(root)

    xml_content = ET.tostring(root, encoding="utf-8").decode("utf-8")

    file_stream = io.BytesIO()
    file_stream.write(xml_content.encode("utf-8"))
    file_stream.seek(0)
    filename = f"{plan.okpo}_{plan.year}.xml"
    return file_stream, "application/xml", filename
    
def export_pdf_single(plan: Plan):
    pass


def export_xlsx_single_before25(plan: Plan):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # ===============================
    # Шрифт и выравнивание
    # ===============================
    regular_font_9 = Font(name="Times New Roman", size=9)
    regular_font_9_italic = Font(name="Times New Roman", size=9, italic=True)
    regular_font_10 = Font(name="Times New Roman", size=10)
    regular_font_10_italic = Font(name="Times New Roman", size=10, italic=True)
    
    regular_font_11 = Font(name="Times New Roman", size=11)
    regular_font_11_italic = Font(name="Times New Roman", size=11, italic=True)
    
    regular_font_13 = Font(name="Times New Roman", size=13)
    bold_font_10 = Font(name="Times New Roman", size=10, bold=True)
    bold_font_11 = Font(name="Times New Roman", size=11, bold=True)
    bold_font_13 = Font(name="Times New Roman", size=13, bold=True)
    
    vertical_text = Alignment(horizontal="center", vertical="bottom", textRotation=90, wrap_text=True)
    top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    bottom = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    bottom_left = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_bottom = Side(border_style="thin", color="000000")
    bottom_border = Border(bottom=thin_bottom)


    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    def set_cell(ws, row_start, col_start, row_end=None, col_end=None, text="", 
                        font=regular_font_10, row_height=None, alignment=left,
                        merge_direction='both'):  # 'horizontal', 'vertical', 'both', 'none'
        """
        merge_direction: 
        - 'horizontal': объединить только по столбцам (row_end игнорируется)
        - 'vertical': объединить только по строкам (col_end игнорируется)
        - 'both': объединить и по строкам и по столбцам
        - 'none': не объединять
        """
        if merge_direction == 'horizontal':
            row_end = row_start
            if col_end is None:
                col_end = col_start
        elif merge_direction == 'vertical':
            col_end = col_start
            if row_end is None:
                row_end = row_start
        elif merge_direction == 'both':
            if row_end is None:
                row_end = row_start
            if col_end is None:
                col_end = col_start
        else:
            row_end = row_start
            col_end = col_start
        
        if row_start != row_end or col_start != col_end:
            ws.merge_cells(start_row=row_start, start_column=col_start, 
                        end_row=row_end, end_column=col_end)
        
        cell = ws.cell(row=row_start, column=col_start)
        cell.value = text
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        
        if row_height is not None:
            for row in range(row_start, row_end + 1):
                ws.row_dimensions[row].height = row_height
        
        return cell
    
    def page_setttings(ws, print_area):
        ws.print_area = print_area      
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

    def get_org_name_by_okpo(okpo):
        dop_org_data = [
            ('Брестское областное управление', '100000001000'),
            ('Витебское областное управление', '200000002000'),
            ('Гомельское областное управление', '300000003000'),
            ('Гродненское областное управление', '400000004000'),
            ('Управление г. Минск', '500000005000'),
            ('Минское областное управление', '600000006000'),
            ('Могилевское областное управление', '700000007000'),
            ('Департамент по энергоэффективности', '800000008000'),
        ]
    
        if not okpo or len(str(okpo)) < 4:
            return ""  
        
        okpo_str = str(okpo)
        if len(okpo_str) >= 4:
            fourth_from_end = okpo_str[-4]  
        else:
            return ""
        
        for name, code in dop_org_data:
            if code.endswith(str(fourth_from_end) + "000"):
                return name
        
            return ""

    def title_xlsx(wb, plan):
        ws = wb.create_sheet("Титульный лист", 0)
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 8.43), ("B", 8.43), ("C", 8.43), ("D", 8.43),
                ("E", 8.43), ("F", 8.43), ("G", 8.43), ("H", 8.43),
                ("I", 8.43), ("J", 8.43), ("K", 8.43), ("L", 8.43),
                ("M", 8.43), ("N", 8.43)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 3:
                ws.row_dimensions[row].height = 12
            elif row == 5:
                ws.row_dimensions[row].height = 32.25
            elif row == 7:
                ws.row_dimensions[row].height = 4
            elif row == 8:
                ws.row_dimensions[row].height = 12
            elif row == 9:
                ws.row_dimensions[row].height = 22.5
            elif row == 16:
                ws.row_dimensions[row].height = 17.5
            elif row == 17:
                ws.row_dimensions[row].height = 19.5
            elif row == 20:
                ws.row_dimensions[row].height = 16.5
            else:
                ws.row_dimensions[row].height = 15

        # ===============================
        # Блоки текста
        # ===============================
        
        def title_first_sign():
            ws.merge_cells("B1:D1")
            ws["B1"].value = "Согласовано".upper()
            ws["B1"].font = bold_font_11

            ws.merge_cells("B2:D2")
            ws["B2"].value = "_______________________"
            ws["B2"].font = bold_font_11
            
            ws.merge_cells("B3:D3")
            ws["B3"].value = "(должность)"
            ws["B3"].font = regular_font_9
            ws["B3"].alignment = center
                    
            ws.merge_cells("B4:F4")
            ws["B4"].value = "_______________ областного (городского)"
            ws["B4"].font = regular_font_11
            ws["B4"].alignment = left
                            
                        
            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"{org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"областное (городское) управление по надзору за рациональным использованием ТЭР"
                        
            ws.merge_cells("B5:F5")
            ws["B5"].value = org_text
            ws["B5"].font = regular_font_11
            ws["B5"].alignment = left
                    
            ws.merge_cells("B6:D6")
            ws["B6"].value = "подписано ЭЦП"
            ws["B6"].font = regular_font_9_italic
            ws["B6"].alignment = center
            
            ws.merge_cells("B7:D7")
            ws["B7"].value = "_______________________"
            ws["B7"].font = bold_font_11
            
            ws.merge_cells("B8:D8")
            ws["B8"].value = "(подпись, инициалы и фамилия)"
            ws["B8"].font = regular_font_9
            ws["B8"].alignment = left
            
            ws.merge_cells("B9:E9")
            ws["B9"].value = "«___» ____________ 20__ г."
            ws["B9"].font = regular_font_11
            ws["B9"].alignment = left

        def title_second_sign():
            ws.merge_cells("K1:M1")
            ws["K1"].value = "Утверждаю".upper()
            ws["K1"].font = bold_font_11

            ws.merge_cells("K2:M2")
            ws["K2"].value = "_______________________"
            ws["K2"].font = bold_font_11
            
            ws.merge_cells("K3:M3")
            ws["K3"].value = "(должность)"
            ws["K3"].font = regular_font_9
            ws["K3"].alignment = center
            
                    
            ws.merge_cells("K4:M4")
            ws["K4"].value = "_______________________"
            ws["K4"].font = regular_font_11
            ws["K4"].alignment = left
                            
            ws.merge_cells("K5:M5")
            ws["K5"].value = "(министерство, концерн, государственный комитет)"
            ws["K5"].font = regular_font_11
            ws["K5"].alignment = left
            
                    
            ws.merge_cells("K6:M6")
            ws["K6"].value = "подписано ЭЦП"
            ws["K6"].font = regular_font_9_italic
            ws["K6"].alignment = center
            
            ws.merge_cells("K7:M7")
            ws["K7"].value = "_______________________"
            ws["K7"].font = bold_font_11
            
            ws.merge_cells("K8:M8")
            ws["K8"].value = "(подпись, инициалы и фамилия)"
            ws["K8"].font = regular_font_9
            ws["K8"].alignment = left
            
            ws.merge_cells("K9:N9")
            ws["K9"].value = "«___» ____________ 20__ г."
            ws["K9"].font = regular_font_11
            ws["K9"].alignment = left
        
        title_first_sign()
        title_second_sign()
        
        ws.merge_cells("A14:O15")
        ws["A14"].value = "ПЛАН МЕРОПРИЯТИЙ ПО ЭНЕРГОСБЕРЕЖЕНИЮ"
        ws["A14"].font = bold_font_13
        ws["A14"].alignment = center
                                     
        ws.merge_cells("B16:N17")
        ws["B16"].value = f"{plan.organization.name}"
        ws["B16"].font = regular_font_13
        ws["B16"].alignment = center
        
        for col in range(2, 15):  # B=2, N=15
            ws.cell(row=17, column=col).border = bottom_border
        
        ws.merge_cells("D18:L18")
        ws["D18"].value = "(наименование юридического лица)"
        ws["D18"].font = regular_font_13
        ws["D18"].alignment = center     
                
        ws.merge_cells("B20:N20")
        ws["B20"].value = f"на {plan.year} год".upper()
        ws["B20"].font = bold_font_13
        ws["B20"].alignment = center
        
        ws.merge_cells("B25:D25")
        ws["B25"].value = "Целевые показатели:"
        ws["B25"].font = bold_font_11
        ws["B25"].alignment = center    
        
        ws.merge_cells("E25:H25")
        ws["E25"].value = "энергосбережения"
        ws["E25"].font = bold_font_11
        ws["E25"].alignment = left 
               
        ws.merge_cells("E26:H26")
        ws["E26"].value = "по экономии ТЭР"
        ws["E26"].font = bold_font_11
        ws["E26"].alignment = left   
             
        ws.merge_cells("E27:H27")
        ws["E27"].value = "по доле местных ТЭР в КПТ"
        ws["E27"].font = bold_font_11
        ws["E27"].alignment = left 
                    
        ws.merge_cells("E28:H28")
        ws["E28"].value = "по доле ВИЭ в КПТ"
        ws["E28"].font = bold_font_11
        ws["E28"].alignment = left
                    
        ws["I25"].value = "-"
        ws["I25"].font = bold_font_11
        ws["I25"].alignment = center                    
        ws["I26"].value = "-"
        ws["I26"].font = bold_font_11
        ws["I26"].alignment = center                    
        ws["I27"].value = "-"
        ws["I27"].font = bold_font_11
        ws["I27"].alignment = center                    
        ws["I28"].value = "-"
        ws["I28"].font = bold_font_11
        ws["I28"].alignment = center

        ws["J25"].value = f"{plan.energy_saving}"
        ws["J25"].font = bold_font_11
        ws["J25"].alignment = center        
        ws["J26"].value = f"{plan.share_fuel}"
        ws["J26"].font = bold_font_11
        ws["J26"].alignment = center        
        ws["J27"].value = f"{plan.saving_fuel}"
        ws["J27"].font = bold_font_11
        ws["J27"].alignment = center        
        ws["J28"].value = f"{plan.share_energy}"
        ws["J28"].font = bold_font_11
        ws["J28"].alignment = center       

        ws["K25"].value = "%"
        ws["K25"].font = bold_font_11
        ws["K25"].alignment = center        
        ws["K26"].value = "т у.т."
        ws["K26"].font = bold_font_11
        ws["K26"].alignment = center        
        ws["K27"].value = "%"
        ws["K27"].font = bold_font_11
        ws["K27"].alignment = center        
        ws["K28"].value = "%"
        ws["K28"].font = bold_font_11
        ws["K28"].alignment = center
                     
        page_setttings(ws, print_area = "A1:O32")
        
        return ws
 
    def first_half_xlsx(wb, plan):
        ws = wb.create_sheet("Часть 1")
    
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 5.43), ("B", 58), ("C", 14), ("D", 10),
                ("E", 10), ("F", 10), ("G", 18.29)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21
            elif row == 2:
                ws.row_dimensions[row].height = 12
            elif row == 3:
                ws.row_dimensions[row].height = 85.5
            elif row == 4:
                ws.row_dimensions[row].height = 34
            else:
                ws.row_dimensions[row].height = 15
        
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        ws.merge_cells("A1:G1")
        ws["A1"].value = "Часть 1. Показатели использования топливно-энергетических ресурсов"
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center
        
        ws.merge_cells("A2:G2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        headers = [
            "№ п/п", 
            "Основные показатели использования ТЭР", 
            "Единица измерения", 
            f"{plan.year - 1} г. отчет", 
            f"{plan.year} г. оценка", 
            f"{plan.year + 1} г. прогноз", 
            "Изменение ТЭР прогнозного года к предыдущему (увеличение + , снижение - )"
        ]
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = bold_font_11
            cell.alignment = center
            cell.border = thin_border
        
        previous_group = None
        row_index = 3

        for usage in sorted(plan.indicators_usage, 
            key=lambda u: (
                (0, u.indicator.Group) if u.indicator.Group is not None else (1, float('-inf')),
                (0, u.indicator.RowN) if u.indicator.RowN is not None else (1, float('-inf'))
            )
        ):

            if usage.indicator.Group != previous_group:
                if usage.indicator.Group is not None:
                    try:
                        group_float = float(usage.indicator.Group)
                        if group_float.is_integer():
                            group_value = int(group_float)
                        else:
                            group_value = usage.indicator.Group
                    except (ValueError, TypeError):
                        group_value = usage.indicator.Group
                else:
                    group_value = ""
            else:
                group_value = ""
            
            previous_group = usage.indicator.Group
            
            row = [
                group_value, 
                usage.indicator.name if usage.indicator.name else "-",
                usage.indicator.unit.name,
                (usage.QYearPrev or 0),
                (usage.QYearCurr or 0),
                (usage.QYearNext or 0),
                (usage.QYearNext or 0) - (usage.QYearCurr or 0),
            ]
            ws.append(row)
            
            row_index += 1
            
            if group_value in [5, 8]:
                ws.row_dimensions[row_index].height = 34
                
            for col in range(1, len(row)+1):
                cell = ws.cell(row=row_index, column=col)
                if group_value != "":
                    cell.font = bold_font_11
                else:
                    cell.font = regular_font_11
                
                cell.alignment = center if col != 2 else left
                cell.border = thin_border
                
                if col == 1:
                    if group_value != "":
                        if isinstance(group_value, (int, float)):
                            try:
                                num = float(group_value)
                                if num.is_integer():
                                    cell.number_format = '0'
                                else:
                                    cell.number_format = '0.0'
                            except:
                                cell.number_format = '@'
                        else:
                            cell.number_format = '@'
                    else:
                        cell.number_format = '@'
                elif col in [4, 5, 6, 7]:
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '@'
        
        def signatures_indicators_xlsx(ws, start_row, plan):
            
            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"_______________ {org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"_______________ областное (городское) управление по надзору за рациональным использованием ТЭР"
            
            set_cell(ws, row_start=start_row+2, col_start=2, row_end=start_row+3, col_end=2, 
                    text=org_text, font=regular_font_10)
            
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=2, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=bottom_left)   
              
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=2, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=2, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=bottom_left) 
             
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=2, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+2, col_start=5, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')           
                                   
            set_cell(ws, row_start=start_row+3, col_start=5, col_end=7, 
                    text="(юридическое лицо)", merge_direction='horizontal', font = regular_font_9, alignment=center)      
                     
            set_cell(ws, row_start=start_row+4, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+5, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+7, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
                         
            set_cell(ws, row_start=start_row+9, col_start=5, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')               
                         
            set_cell(ws, row_start=start_row+10, col_start=5, col_end=7, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font = regular_font_9, row_height = 11)  
                                 
            set_cell(ws, row_start=start_row+11, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+12, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+13, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+14, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                 
        signatures_indicators_xlsx(ws, row_index + 1, plan)

        page_setttings(ws, print_area = "A1:G75")

        return ws

    def second_half_xlsx(wb, plan):
        ws = wb.create_sheet("Часть 2")
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 4), 
                ("B", 9), 
                ("C", 15), 
                ("D", 6),
                ("E", 6.75), 
                ("F", 6.43), 
                ("G", 6.86), 
                ("H", 6.43), 
                ("I", 9.43), 
                ("J", 5), 
                ("K", 7.86), 
                ("L", 7.43), 
                ("M", 6.57), 
                ("N", 6.57), 
                ("O", 6), 
                ("P", 7.14), 
                ("Q", 6.71), 
                ("R", 7)
                ]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21.75
            elif row == 4:
                ws.row_dimensions[row].height = 23.25
            elif row == 5:
                ws.row_dimensions[row].height = 19.5
            elif row == 6:
                ws.row_dimensions[row].height = 159
            else:
                ws.row_dimensions[row].height = 15

        ws.merge_cells("A1:R1")
        ws["A1"].value = "Часть 2. Мероприятия по экономии топливно-энергетических ресурсов"
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center

        ws.merge_cells("A2:R2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        ws.merge_cells("A3:A6"); ws["A3"].value = "№ п/п"
        ws.merge_cells("B3:B6"); ws["B3"].value = "Код основных направлений энергосбережения"
        ws.merge_cells("C3:C6"); ws["C3"].value = "Наименование мероприятия"
        ws.merge_cells("D3:D6"); ws["D3"].value = "Единицы измерения"
        ws.merge_cells("E3:E6"); ws["E3"].value = "Объем внедрения"
        ws.merge_cells("F3:G5"); ws["F3"].value = "Условно-годовой экономический эффект"
        ws["F6"].value = "т у.т."
        ws["G6"].value = "тыс. руб."
        ws.merge_cells("H3:H6"); ws["H3"].value = "Ожидаемый срок внедрения мероприятия, квартал"
        ws.merge_cells("I3:I6"); ws["I3"].value = "Ожидаемый экономический эффект от внедрения мероприятия в текущем году, т у.т."
        ws.merge_cells("J3:J6"); ws["J3"].value = "Срок окупаемости, лет"
        ws.merge_cells("K3:K6"); ws["K3"].value = "Объем финансирования, тыс. руб."
        ws.merge_cells("L3:R3"); ws["L3"].value = "источники финансирования, руб."
        ws.merge_cells("L4:O4"); ws["L4"].value = "бюджетные"
        ws.merge_cells("L5:L6"); ws["L5"].value = "республиканский бюджет на финансирование госпрограммы"
        ws.merge_cells("M5:M6"); ws["M5"].value = "республиканский бюджет"
        ws.merge_cells("N5:N6"); ws["N5"].value = "местный бюджет"
        ws.merge_cells("O5:O6"); ws["O5"].value = "другие"
        ws.merge_cells("P4:P6"); ws["P4"].value = "собственные средства организации"
        ws.merge_cells("Q4:Q6"); ws["Q4"].value = "кредиты банков, займы"
        ws.merge_cells("R4:R6"); ws["R4"].value = "иные"

        for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=18):
            for cell in row:
                if not cell.value:
                    continue
                if cell.coordinate in ("D3", "E3", "F6", "G6", "J3", "K3", "H3", "I3", "L5", "M5", "N5", "O5", "P4", "Q4", "R4"):
                    cell.alignment = vertical_text
                else:
                    cell.alignment = top
                cell.font = regular_font_10

        row_index = 7
        for col in range(1, 19):
            cell = ws.cell(row=row_index, column=col, value=col)
            cell.alignment = center
            cell.font = regular_font_10

        local_econ_execes = (EconExec.query
            .join(EconMeasure)
            .join(Plan)
            .filter(Plan.id == plan.id, EconExec.is_local == True)
            .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
            .all()
        )

        non_local_econ_execes = (EconExec.query
            .join(EconMeasure)
            .join(Plan)
            .filter(Plan.id == plan.id, EconExec.is_local == False)
            .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
            .all()
        )

        def add_section(title, execs, start_number=1):
            nonlocal row_index
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=18)
            ws.cell(row=row_index, column=1, value=title).font = bold_font_10
            ws.cell(row=row_index, column=1).alignment = center
            sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
            sums = {col: 0 for col in sum_cols}
            for idx, econ in enumerate(execs, start=start_number):
                row_index += 1
                row = [
                    idx,
                    econ.econ_measures.direction.code if econ.econ_measures and econ.econ_measures.direction else "",
                    econ.name if hasattr(econ, "name") else "",
                    econ.econ_measures.direction.unit.name if econ.econ_measures and econ.econ_measures.direction and econ.econ_measures.direction.unit else "",
                    econ.Volume if hasattr(econ, "Volume") else 0,
                    econ.EffTut if hasattr(econ, "EffTut") else 0,
                    econ.EffRub if hasattr(econ, "EffRub") else 0,
                    econ.ExpectedQuarter if hasattr(econ, "ExpectedQuarter") else 0,
                    econ.EffCurrYear if hasattr(econ, "EffCurrYear") else 0,
                    econ.Payback if hasattr(econ, "Payback") else 0,
                    econ.VolumeFin if hasattr(econ, "VolumeFin") else 0,
                    econ.BudgetState if hasattr(econ, "BudgetState") else 0,
                    econ.BudgetRep if hasattr(econ, "BudgetRep") else 0,
                    econ.BudgetLoc if hasattr(econ, "BudgetLoc") else 0,
                    econ.BudgetOther if hasattr(econ, "BudgetOther") else 0,
                    econ.MoneyOwn if hasattr(econ, "MoneyOwn") else 0,
                    econ.MoneyLoan if hasattr(econ, "MoneyLoan") else 0,
                    econ.MoneyOther if hasattr(econ, "MoneyOther") else 0
                ]
                for col in sum_cols:
                    try:
                        sums[col] += float(row[col-1])
                    except (TypeError, ValueError):
                        pass
                ws.append(row)
                for col_idx in range(1, 19):
                    cell = ws.cell(row=row_index, column=col_idx)
                    cell.alignment = left if col_idx == 3 else center
                    cell.font = regular_font_10
                    if col_idx in [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]:
                        cell.number_format = '0.000'
            row_index += 1
            
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value="Итого по разделу:").alignment = left
            ws.cell(row=row_index, column=1).font = regular_font_10_italic
            
            for col in sum_cols:
                cell = ws.cell(row=row_index, column=col, value=sums[col])
                cell.alignment = center
                cell.font = regular_font_10_italic
                cell.number_format = '0.000'
            return start_number + len(execs)

        next_number = add_section("Раздел 2. Мероприятия по экономии топливно-энергетических ресурсов", non_local_econ_execes, 1)
        # add_section("Раздел 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов", local_econ_execes, next_number)

        # row_index += 1
        
        # ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        # ws.cell(row=row_index, column=1, value="Всего по части 2:").font = bold_font_10
        # ws.cell(row=row_index, column=1).alignment = left
        
        # sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
        # total_sums = {col: 0 for col in sum_cols}
        # for row in ws.iter_rows(min_row=6, max_row=row_index, min_col=3, max_col=18):
        #     if row[0].value == "Итого по разделу:":
        #         for col in sum_cols:
        #             val = ws.cell(row=row[0].row, column=col).value
        #             try:
        #                 total_sums[col] += float(val)
        #             except (TypeError, ValueError):
        #                 pass
        # for col in sum_cols:
        #     cell = ws.cell(row=row_index, column=col, value=total_sums[col])
        #     cell.alignment = center
        #     cell.font = regular_font_10_italic
        #     cell.number_format = '0.000'

        quarters = [
            ("      январь–март", "jan_mar"),
            ("      январь–июнь", "jan_jun"),
            ("      январь–сентябрь", "jan_sep"),
            ("      январь–декабрь", "jan_dec")
        ]
        local_totals = get_cumulative_econ_metrics(plan.id, True)
        
        non_local_totals = get_cumulative_econ_metrics(plan.id, False)
        # total_metrics = {
        #     'jan_mar_eff': local_totals['jan_mar']['eff_curr_year'] + non_local_totals['jan_mar']['eff_curr_year'],
        #     'jan_mar_vol': local_totals['jan_mar']['volume_fin'] + non_local_totals['jan_mar']['volume_fin'],
        #     'jan_jun_eff': local_totals['jan_jun']['eff_curr_year'] + non_local_totals['jan_jun']['eff_curr_year'],
        #     'jan_jun_vol': local_totals['jan_jun']['volume_fin'] + non_local_totals['jan_jun']['volume_fin'],
        #     'jan_sep_eff': local_totals['jan_sep']['eff_curr_year'] + non_local_totals['jan_sep']['eff_curr_year'],
        #     'jan_sep_vol': local_totals['jan_sep']['volume_fin'] + non_local_totals['jan_sep']['volume_fin'],
        #     'jan_dec_eff': local_totals['jan_dec']['eff_curr_year'] + non_local_totals['jan_dec']['eff_curr_year'],
        #     'jan_dec_vol': local_totals['jan_dec']['volume_fin'] + non_local_totals['jan_dec']['volume_fin']
        # }
        
        total_metrics = {
            'jan_mar_eff': local_totals['jan_mar']['eff_curr_year'],
            'jan_mar_vol': local_totals['jan_mar']['volume_fin'],
            'jan_jun_eff': local_totals['jan_jun']['eff_curr_year'],
            'jan_jun_vol': local_totals['jan_jun']['volume_fin'],
            'jan_sep_eff': local_totals['jan_sep']['eff_curr_year'],
            'jan_sep_vol': local_totals['jan_sep']['volume_fin'],
            'jan_dec_eff': local_totals['jan_dec']['eff_curr_year'],
            'jan_dec_vol': local_totals['jan_dec']['volume_fin']
        }


        for q_label, q_key in quarters:
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value=q_label).font = regular_font_10
            ws.cell(row=row_index, column=1).alignment = left
            
            for col in range(3, 19):
                c = ws.cell(row=row_index, column=col)
                c.alignment = center
            
            eff_cell = ws.cell(row=row_index, column=9, value=total_metrics[f"{q_key}_eff"])
            eff_cell.number_format = '0.000'
            eff_cell.font = regular_font_10
            
            vol_cell = ws.cell(row=row_index, column=11, value=total_metrics[f"{q_key}_vol"])
            vol_cell.number_format = '0.000'
            vol_cell.font = regular_font_10 

        for row in ws.iter_rows(min_row=3, max_row=row_index, min_col=1, max_col=18):
            for cell in row:
                cell.border = thin_border

        def signatures_second_half_xlsx(ws, start_row, plan):
            big_space = " " * 10  
            text = f"Стоимость 1 т у.т. принята равной{big_space}долларов, при курсе{big_space}руб."

            set_cell(ws, row_start=start_row, col_start=2, col_end=9, 
                    text=text, font=regular_font_10_italic, row_height=None)

            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"_______________ {org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"_______________ областное (городское) управление по надзору за рациональным использованием ТЭР"

            set_cell(ws, row_start=start_row+2, col_start=2, row_end=start_row+3, col_end=9, text=org_text)
      
      
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=3, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)   
              
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=3, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=3, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=center) 
             
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=3, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+2, col_start=12, col_end=17, 
                    text="__________________________________", merge_direction='horizontal')           
                                   
            set_cell(ws, row_start=start_row+3, col_start=12, col_end=16, 
                    text="(юридическое лицо)", merge_direction='horizontal', font = regular_font_9, alignment=center)      
                     
            set_cell(ws, row_start=start_row+4, col_start=12, col_end=15, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+5, col_start=12, col_end=15, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=12, col_end=15, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+7, col_start=12, col_end=15, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
                         
            set_cell(ws, row_start=start_row+9, col_start=12, col_end=17, 
                    text="__________________________________", merge_direction='horizontal')               
                         
            set_cell(ws, row_start=start_row+10, col_start=12, col_end=17, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font = regular_font_9, row_height = 11)  
                                 
            set_cell(ws, row_start=start_row+11, col_start=12, col_end=15, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+12, col_start=12, col_end=15, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+13, col_start=12, col_end=15, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+14, col_start=12, col_end=15, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
              
        signatures_second_half_xlsx(ws, row_index + 1, plan)
        page_setttings(ws, print_area="A1:R32")
        
        return ws

    def third_half_xlsx(wb, plan):
        ws = wb.create_sheet("Часть 3")
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 4), 
                ("B", 9), 
                ("C", 15), 
                ("D", 6),
                ("E", 6.75), 
                ("F", 6.43), 
                ("G", 6.86), 
                ("H", 6.43), 
                ("I", 9.43), 
                ("J", 5), 
                ("K", 7.86), 
                ("L", 7.43), 
                ("M", 6.57), 
                ("N", 6.57), 
                ("O", 6), 
                ("P", 7.14), 
                ("Q", 6.71), 
                ("R", 7)
                ]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21.75
            elif row == 4:
                ws.row_dimensions[row].height = 23.25
            elif row == 5:
                ws.row_dimensions[row].height = 26
            elif row == 6:
                ws.row_dimensions[row].height = 159
            else:
                ws.row_dimensions[row].height = 15

        ws.merge_cells("A1:R1")
        ws["A1"].value = "Часть 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов"
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center

        ws.merge_cells("A2:R2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        ws.merge_cells("A3:A6"); ws["A3"].value = "№ п/п"
        ws.merge_cells("B3:B6"); ws["B3"].value = "Код основных направлений энергосбережения"
        ws.merge_cells("C3:C6"); ws["C3"].value = "Наименование мероприятия"
        ws.merge_cells("D3:D6"); ws["D3"].value = "Единицы измерения"
        ws.merge_cells("E3:E6"); ws["E3"].value = "Объем внедрения"
        ws.merge_cells("F3:G5"); ws["F3"].value = "Условно-годовое увеличение использования местных ТЭР "
        ws["F6"].value = "т у.т."
        ws["G6"].value = "тыс. руб."
        ws.merge_cells("H3:H6"); ws["H3"].value = "Ожидаемый срок внедрения мероприятия, квартал"
        ws.merge_cells("I3:I6"); ws["I3"].value = "Ожидаемый экономический эффект от внедрения мероприятия в текущем году, т у.т."
        ws.merge_cells("J3:J6"); ws["J3"].value = "Срок окупаемости, лет"
        ws.merge_cells("K3:K6"); ws["K3"].value = "Объем финансирования, тыс. руб."
        ws.merge_cells("L3:R3"); ws["L3"].value = "источники финансирования, руб."
        ws.merge_cells("L4:O4"); ws["L4"].value = "бюджетные"
        ws.merge_cells("L5:L6"); ws["L5"].value = "республиканский бюджет на финансирование госпрограммы"
        ws.merge_cells("M5:M6"); ws["M5"].value = "республиканский бюджет"
        ws.merge_cells("N5:N6"); ws["N5"].value = "местный бюджет"
        ws.merge_cells("O5:O6"); ws["O5"].value = "другие"
        ws.merge_cells("P4:P6"); ws["P4"].value = "собственные средства организации"
        ws.merge_cells("Q4:Q6"); ws["Q4"].value = "кредиты банков, займы"
        ws.merge_cells("R4:R6"); ws["R4"].value = "иные"

        for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=18):
            for cell in row:
                if not cell.value:
                    continue
                if cell.coordinate in ("D3", "E3", "F6", "G6", "J3", "K3", "H3", "I3", "L5", "M5", "N5", "O5", "P4", "Q4", "R4"):
                    cell.alignment = vertical_text
                else:
                    cell.alignment = top
                cell.font = regular_font_10

        row_index = 7
        for col in range(1, 19):
            cell = ws.cell(row=row_index, column=col, value=col)
            cell.alignment = center
            cell.font = regular_font_10

        local_econ_execes = (EconExec.query
            .join(EconMeasure)
            .join(Plan)
            .filter(Plan.id == plan.id, EconExec.is_local == True)
            .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
            .all()
        )

        non_local_econ_execes = (EconExec.query
            .join(EconMeasure)
            .join(Plan)
            .filter(Plan.id == plan.id, EconExec.is_local == False)
            .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
            .all()
        )

        def add_section(title, execs, start_number=1):
            nonlocal row_index
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=18)
            ws.cell(row=row_index, column=1, value=title).font = bold_font_10
            ws.cell(row=row_index, column=1).alignment = center
            sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
            sums = {col: 0 for col in sum_cols}
            for idx, econ in enumerate(execs, start=start_number):
                row_index += 1
                row = [
                    idx,
                    econ.econ_measures.direction.code if econ.econ_measures and econ.econ_measures.direction else "",
                    econ.name if hasattr(econ, "name") else "",
                    econ.econ_measures.direction.unit.name if econ.econ_measures and econ.econ_measures.direction and econ.econ_measures.direction.unit else "",
                    econ.Volume if hasattr(econ, "Volume") else 0,
                    econ.EffTut if hasattr(econ, "EffTut") else 0,
                    econ.EffRub if hasattr(econ, "EffRub") else 0,
                    econ.ExpectedQuarter if hasattr(econ, "ExpectedQuarter") else 0,
                    econ.EffCurrYear if hasattr(econ, "EffCurrYear") else 0,
                    econ.Payback if hasattr(econ, "Payback") else 0,
                    econ.VolumeFin if hasattr(econ, "VolumeFin") else 0,
                    econ.BudgetState if hasattr(econ, "BudgetState") else 0,
                    econ.BudgetRep if hasattr(econ, "BudgetRep") else 0,
                    econ.BudgetLoc if hasattr(econ, "BudgetLoc") else 0,
                    econ.BudgetOther if hasattr(econ, "BudgetOther") else 0,
                    econ.MoneyOwn if hasattr(econ, "MoneyOwn") else 0,
                    econ.MoneyLoan if hasattr(econ, "MoneyLoan") else 0,
                    econ.MoneyOther if hasattr(econ, "MoneyOther") else 0
                ]
                for col in sum_cols:
                    try:
                        sums[col] += float(row[col-1])
                    except (TypeError, ValueError):
                        pass
                ws.append(row)
                for col_idx in range(1, 19):
                    cell = ws.cell(row=row_index, column=col_idx)
                    cell.alignment = left if col_idx == 3 else center
                    cell.font = regular_font_10
                    if col_idx in [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]:
                        cell.number_format = '0.000'
            row_index += 1
            
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value="Итого по разделу:").alignment = left
            ws.cell(row=row_index, column=1).font = regular_font_10_italic
            
            for col in sum_cols:
                cell = ws.cell(row=row_index, column=col, value=sums[col])
                cell.alignment = center
                cell.font = regular_font_10_italic
                cell.number_format = '0.000'
            return start_number + len(execs)

        add_section("Раздел 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов", local_econ_execes, 1)

        # row_index += 1
        
        # ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        # ws.cell(row=row_index, column=1, value="Всего по части 2:").font = bold_font_10
        # ws.cell(row=row_index, column=1).alignment = left
        
        # sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
        # total_sums = {col: 0 for col in sum_cols}
        # for row in ws.iter_rows(min_row=6, max_row=row_index, min_col=3, max_col=18):
        #     if row[0].value == "Итого по разделу:":
        #         for col in sum_cols:
        #             val = ws.cell(row=row[0].row, column=col).value
        #             try:
        #                 total_sums[col] += float(val)
        #             except (TypeError, ValueError):
        #                 pass
        # for col in sum_cols:
        #     cell = ws.cell(row=row_index, column=col, value=total_sums[col])
        #     cell.alignment = center
        #     cell.font = regular_font_10_italic
        #     cell.number_format = '0.000'

        quarters = [
            ("      январь–март", "jan_mar"),
            ("      январь–июнь", "jan_jun"),
            ("      январь–сентябрь", "jan_sep"),
            ("      январь–декабрь", "jan_dec")
        ]
        local_totals = get_cumulative_econ_metrics(plan.id, True)
        non_local_totals = get_cumulative_econ_metrics(plan.id, False)
        
        total_metrics = {
            'jan_mar_eff': non_local_totals['jan_mar']['eff_curr_year'],
            'jan_mar_vol': non_local_totals['jan_mar']['volume_fin'],
            'jan_jun_eff': non_local_totals['jan_jun']['eff_curr_year'],
            'jan_jun_vol': non_local_totals['jan_jun']['volume_fin'],
            'jan_sep_eff': non_local_totals['jan_sep']['eff_curr_year'],
            'jan_sep_vol': non_local_totals['jan_sep']['volume_fin'],
            'jan_dec_eff': non_local_totals['jan_dec']['eff_curr_year'],
            'jan_dec_vol': non_local_totals['jan_dec']['volume_fin']
        }

        for q_label, q_key in quarters:
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value=q_label).font = regular_font_10
            ws.cell(row=row_index, column=1).alignment = left
            
            for col in range(3, 19):
                c = ws.cell(row=row_index, column=col)
                c.alignment = center
            
            eff_cell = ws.cell(row=row_index, column=9, value=total_metrics[f"{q_key}_eff"])
            eff_cell.number_format = '0.000'
            eff_cell.font = regular_font_10
            
            vol_cell = ws.cell(row=row_index, column=11, value=total_metrics[f"{q_key}_vol"])
            vol_cell.number_format = '0.000'
            vol_cell.font = regular_font_10 

        for row in ws.iter_rows(min_row=3, max_row=row_index, min_col=1, max_col=18):
            for cell in row:
                cell.border = thin_border

        def signatures_third_half_xlsx(ws, start_row, plan):
            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"_______________ {org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"_______________ областное (городское) управление по надзору за рациональным использованием ТЭР"

            set_cell(ws, row_start=start_row+2, col_start=2, row_end=start_row+3, col_end=9, text=org_text)
    
    
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=3, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)   
            
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=3, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                    
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=3, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=center) 
            
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=3, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+2, col_start=12, col_end=17, 
                    text="__________________________________", merge_direction='horizontal')           
                                
            set_cell(ws, row_start=start_row+3, col_start=12, col_end=16, 
                    text="(юридическое лицо)", merge_direction='horizontal', font = regular_font_9, alignment=center)      
                    
            set_cell(ws, row_start=start_row+4, col_start=12, col_end=15, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                    
            set_cell(ws, row_start=start_row+5, col_start=12, col_end=15, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                    
            set_cell(ws, row_start=start_row+6, col_start=12, col_end=15, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                        
            set_cell(ws, row_start=start_row+7, col_start=12, col_end=15, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
                        
            set_cell(ws, row_start=start_row+9, col_start=12, col_end=17, 
                    text="__________________________________", merge_direction='horizontal')               
                        
            set_cell(ws, row_start=start_row+10, col_start=12, col_end=17, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font = regular_font_9, row_height = 11)  
                                
            set_cell(ws, row_start=start_row+11, col_start=12, col_end=15, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                    
            set_cell(ws, row_start=start_row+12, col_start=12, col_end=15, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                    
            set_cell(ws, row_start=start_row+13, col_start=12, col_end=15, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                        
            set_cell(ws, row_start=start_row+14, col_start=12, col_end=15, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
            
        signatures_third_half_xlsx(ws, row_index + 1, plan)
        page_setttings(ws, print_area="A1:R32")
        
        return ws

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_title = title_xlsx(wb, plan)
    ws_firstHalf = first_half_xlsx(wb, plan)
    ws_secondHalf = second_half_xlsx(wb, plan)
    ws_thirdHalf = third_half_xlsx(wb, plan)
    
    wb.active = wb.index(ws_title)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"{plan.organization.okpo}_{plan.year}.xlsx"

    return (
        file_stream,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename
    )
 
def export_xlsx_single_over25(plan: Plan):
    pass

def export_xlsx_single_region(plan: Plan):
    pass
 
def export_xlsx_single_ministry(plan: Plan):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # ===============================
    # Шрифт и выравнивание
    # ===============================
    regular_font_9 = Font(name="Times New Roman", size=9)
    regular_font_9_italic = Font(name="Times New Roman", size=9, italic=True)
    regular_font_10 = Font(name="Times New Roman", size=10)
    regular_font_10_italic = Font(name="Times New Roman", size=10, italic=True)
    
    regular_font_11 = Font(name="Times New Roman", size=11)
    regular_font_11_italic = Font(name="Times New Roman", size=11, italic=True)
    
    regular_font_13 = Font(name="Times New Roman", size=13)
    bold_font_10 = Font(name="Times New Roman", size=10, bold=True)
    bold_font_11 = Font(name="Times New Roman", size=11, bold=True)
    bold_font_13 = Font(name="Times New Roman", size=13, bold=True)
    
    vertical_text = Alignment(horizontal="center", vertical="bottom", textRotation=90, wrap_text=True)
    top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    bottom = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    bottom_left = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_bottom = Side(border_style="thin", color="000000")
    bottom_border = Border(bottom=thin_bottom)


    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    def set_cell(ws, row_start, col_start, row_end=None, col_end=None, text="", 
                        font=regular_font_10, row_height=None, alignment=left,
                        merge_direction='both'):  # 'horizontal', 'vertical', 'both', 'none'
        """
        merge_direction: 
        - 'horizontal': объединить только по столбцам (row_end игнорируется)
        - 'vertical': объединить только по строкам (col_end игнорируется)
        - 'both': объединить и по строкам и по столбцам
        - 'none': не объединять
        """
        if merge_direction == 'horizontal':
            row_end = row_start
            if col_end is None:
                col_end = col_start
        elif merge_direction == 'vertical':
            col_end = col_start
            if row_end is None:
                row_end = row_start
        elif merge_direction == 'both':
            if row_end is None:
                row_end = row_start
            if col_end is None:
                col_end = col_start
        else:
            row_end = row_start
            col_end = col_start
        
        if row_start != row_end or col_start != col_end:
            ws.merge_cells(start_row=row_start, start_column=col_start, 
                        end_row=row_end, end_column=col_end)
        
        cell = ws.cell(row=row_start, column=col_start)
        cell.value = text
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        
        if row_height is not None:
            for row in range(row_start, row_end + 1):
                ws.row_dimensions[row].height = row_height
        
        return cell
    
    def page_setttings(ws, print_area):
        ws.print_area = print_area      
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3


    def title_xlsx(wb, plan):
        ws = wb.create_sheet("Титульный лист", 0)
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 8.43), ("B", 8.43), ("C", 8.43), ("D", 8.43),
                ("E", 8.43), ("F", 8.43), ("G", 8.43), ("H", 8.43),
                ("I", 8.43), ("J", 8.43), ("K", 8.43), ("L", 8.43),
                ("M", 8.43), ("N", 8.43)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 3:
                ws.row_dimensions[row].height = 12
            elif row == 5:
                ws.row_dimensions[row].height = 32.25
            elif row == 7:
                ws.row_dimensions[row].height = 4
            elif row == 8:
                ws.row_dimensions[row].height = 12
            elif row == 9:
                ws.row_dimensions[row].height = 22.5
            elif row == 16:
                ws.row_dimensions[row].height = 17.5
            elif row == 17:
                ws.row_dimensions[row].height = 19.5
            elif row == 20:
                ws.row_dimensions[row].height = 16.5
            else:
                ws.row_dimensions[row].height = 15

        # ===============================
        # Блоки текста
        # ===============================
        
        def title_first_sign():
            ws.merge_cells("B1:D1")
            ws["B1"].value = "Согласовано".upper()
            ws["B1"].font = bold_font_11

            ws.merge_cells("B2:D2")
            ws["B2"].value = "_______________________"
            ws["B2"].font = bold_font_11
            
            ws.merge_cells("B3:D3")
            ws["B3"].value = "(должность)"
            ws["B3"].font = regular_font_9
            ws["B3"].alignment = center
                    
            ws.merge_cells("B4:F4")
            ws["B4"].value = "Департамента по энергоэффективности"
            ws["B4"].font = regular_font_11
            ws["B4"].alignment = left
                            
      
            ws.merge_cells("B5:F5")
            ws["B5"].value = "Госстандарта"
            ws["B5"].font = regular_font_11
            ws["B5"].alignment = left
                    
            ws.merge_cells("B6:D6")
            ws["B6"].value = "подписано ЭЦП"
            ws["B6"].font = regular_font_9_italic
            ws["B6"].alignment = center
            
            ws.merge_cells("B7:D7")
            ws["B7"].value = "_______________________"
            ws["B7"].font = bold_font_11
            
            ws.merge_cells("B8:D8")
            ws["B8"].value = "(подпись, инициалы и фамилия)"
            ws["B8"].font = regular_font_9
            ws["B8"].alignment = left
            
            ws.merge_cells("B9:E9")
            ws["B9"].value = "«___» ____________ 20__ г."
            ws["B9"].font = regular_font_11
            ws["B9"].alignment = left

        def title_second_sign():
            ws.merge_cells("K1:M1")
            ws["K1"].value = "Утверждаю".upper()
            ws["K1"].font = bold_font_11

            ws.merge_cells("K2:M2")
            ws["K2"].value = "_______________________"
            ws["K2"].font = bold_font_11
            
            ws.merge_cells("K3:M3")
            ws["K3"].value = "(должность)"
            ws["K3"].font = regular_font_9
            ws["K3"].alignment = center
            
                    
            ws.merge_cells("K4:M4")
            ws["K4"].value = "_______________________"
            ws["K4"].font = regular_font_11
            ws["K4"].alignment = left
                            
            ws.merge_cells("K5:M5")
            ws["K5"].value = "(министерство, концерн, государственный комитет)"
            ws["K5"].font = regular_font_11
            ws["K5"].alignment = left
            
                    
            ws.merge_cells("K6:M6")
            ws["K6"].value = "подписано ЭЦП"
            ws["K6"].font = regular_font_9_italic
            ws["K6"].alignment = center
            
            ws.merge_cells("K7:M7")
            ws["K7"].value = "_______________________"
            ws["K7"].font = bold_font_11
            
            ws.merge_cells("K8:M8")
            ws["K8"].value = "(подпись, инициалы и фамилия)"
            ws["K8"].font = regular_font_9
            ws["K8"].alignment = left
            
            ws.merge_cells("K9:N9")
            ws["K9"].value = "«___» ____________ 20__ г."
            ws["K9"].font = regular_font_11
            ws["K9"].alignment = left
        
        title_first_sign()
        title_second_sign()
        
        ws.merge_cells("A13:O15")
        ws["A14"].value = "ПЕРЕЧЕНЬ МЕРОПРИЯТИЙ, НАПРАВЛЕННЫХ НА ДОСТИЖЕНИЕ ЦЕЛЕВЫХ ПОКАЗАТЕЛЕЙ В СФЕРЕ ЭНЕРГОСБЕРЕЖЕНИЯ ГОСУДАРСТВЕННОЙ ПРОГРАММЫ «УСТОЙЧИВАЯ ЭНЕРГЕТИКА И ЭНЕРГОЭФФЕКТИВНОСТЬ» "
        ws["A14"].font = bold_font_13
        ws["A14"].alignment = center
                                     
        ws.merge_cells("B16:N17")
        ws["B16"].value = f"{plan.ministry.name}"
        ws["B16"].font = regular_font_13
        ws["B16"].alignment = center
        
        for col in range(2, 15):  # B=2, N=15
            ws.cell(row=17, column=col).border = bottom_border
        
        ws.merge_cells("D18:L18")
        ws["D18"].value = "(министерство, концерн, государственный комитет)"
        ws["D18"].font = regular_font_13
        ws["D18"].alignment = center     
                
        ws.merge_cells("B20:N20")
        ws["B20"].value = f"на {plan.year} год".upper()
        ws["B20"].font = bold_font_13
        ws["B20"].alignment = center
        
        ws.merge_cells("B25:D25")
        ws["B25"].value = "Целевые показатели:"
        ws["B25"].font = bold_font_11
        ws["B25"].alignment = center    
        
        ws.merge_cells("E25:H25")
        ws["E25"].value = "энергосбережения"
        ws["E25"].font = bold_font_11
        ws["E25"].alignment = left 
               
        ws.merge_cells("E26:H26")
        ws["E26"].value = "по экономии ТЭР"
        ws["E26"].font = bold_font_11
        ws["E26"].alignment = left   
             
        ws.merge_cells("E27:H27")
        ws["E27"].value = "по доле местных ТЭР в КПТ"
        ws["E27"].font = bold_font_11
        ws["E27"].alignment = left 
                    
        ws.merge_cells("E28:H28")
        ws["E28"].value = "по доле ВИЭ в КПТ"
        ws["E28"].font = bold_font_11
        ws["E28"].alignment = left
                    
        ws["I25"].value = "-"
        ws["I25"].font = bold_font_11
        ws["I25"].alignment = center                    
        ws["I26"].value = "-"
        ws["I26"].font = bold_font_11
        ws["I26"].alignment = center                    
        ws["I27"].value = "-"
        ws["I27"].font = bold_font_11
        ws["I27"].alignment = center                    
        ws["I28"].value = "-"
        ws["I28"].font = bold_font_11
        ws["I28"].alignment = center

        ws["J25"].value = f"{plan.energy_saving}"
        ws["J25"].font = bold_font_11
        ws["J25"].alignment = center        
        ws["J26"].value = f"{plan.share_fuel}"
        ws["J26"].font = bold_font_11
        ws["J26"].alignment = center        
        ws["J27"].value = f"{plan.saving_fuel}"
        ws["J27"].font = bold_font_11
        ws["J27"].alignment = center        
        ws["J28"].value = f"{plan.share_energy}"
        ws["J28"].font = bold_font_11
        ws["J28"].alignment = center       

        ws["K25"].value = "%"
        ws["K25"].font = bold_font_11
        ws["K25"].alignment = center        
        ws["K26"].value = "т у.т."
        ws["K26"].font = bold_font_11
        ws["K26"].alignment = center        
        ws["K27"].value = "%"
        ws["K27"].font = bold_font_11
        ws["K27"].alignment = center        
        ws["K28"].value = "%"
        ws["K28"].font = bold_font_11
        ws["K28"].alignment = center
                     
        page_setttings(ws, print_area = "A1:O32")
        
        return ws
 
    def first_half_xlsx(wb, plan):
        ws = wb.create_sheet("Часть 1")
    
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 5.43), ("B", 58), ("C", 14), ("D", 10),
                ("E", 10), ("F", 10), ("G", 18.29)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21
            elif row == 2:
                ws.row_dimensions[row].height = 12
            elif row == 3:
                ws.row_dimensions[row].height = 85.5
            elif row == 4:
                ws.row_dimensions[row].height = 34
            else:
                ws.row_dimensions[row].height = 15
        
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        ws.merge_cells("A1:G1")
        ws["A1"].value = "Часть 1. Показатели использования топливно-энергетических ресурсов"
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center
        
        ws.merge_cells("A2:G2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        headers = [
            "№ п/п", 
            "Основные показатели использования ТЭР", 
            "Единица измерения", 
            f"{plan.year - 1} г. отчет", 
            f"{plan.year} г. оценка", 
            f"{plan.year + 1} г. прогноз", 
            "Изменение ТЭР прогнозного года к предыдущему (увеличение + , снижение - )"
        ]
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = bold_font_11
            cell.alignment = center
            cell.border = thin_border
        
        previous_group = None
        row_index = 3

        for usage in sorted(plan.indicators_usage, 
            key=lambda u: (
                (0, u.indicator.Group) if u.indicator.Group is not None else (1, float('-inf')),
                (0, u.indicator.RowN) if u.indicator.RowN is not None else (1, float('-inf'))
            )
        ):

            if usage.indicator.Group != previous_group:
                if usage.indicator.Group is not None:
                    try:
                        group_float = float(usage.indicator.Group)
                        if group_float.is_integer():
                            group_value = int(group_float)
                        else:
                            group_value = usage.indicator.Group
                    except (ValueError, TypeError):
                        group_value = usage.indicator.Group
                else:
                    group_value = ""
            else:
                group_value = ""
            
            previous_group = usage.indicator.Group
            
            row = [
                group_value, 
                usage.indicator.name if usage.indicator.name else "-",
                usage.indicator.unit.name,
                (usage.QYearPrev or 0),
                (usage.QYearCurr or 0),
                (usage.QYearNext or 0),
                (usage.QYearNext or 0) - (usage.QYearCurr or 0),
            ]
            ws.append(row)
            
            row_index += 1
            
            if group_value in [5, 8]:
                ws.row_dimensions[row_index].height = 34
                
            for col in range(1, len(row)+1):
                cell = ws.cell(row=row_index, column=col)
                if group_value != "":
                    cell.font = bold_font_11
                else:
                    cell.font = regular_font_11
                
                cell.alignment = center if col != 2 else left
                cell.border = thin_border
                
                if col == 1:
                    if group_value != "":
                        if isinstance(group_value, (int, float)):
                            try:
                                num = float(group_value)
                                if num.is_integer():
                                    cell.number_format = '0'
                                else:
                                    cell.number_format = '0.0'
                            except:
                                cell.number_format = '@'
                        else:
                            cell.number_format = '@'
                    else:
                        cell.number_format = '@'
                elif col in [4, 5, 6, 7]:
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '@'
        
        def signatures_indicators_xlsx(ws, start_row, plan):
            
            
            set_cell(ws, row_start=start_row+2, col_start=2, row_end=start_row+3, col_end=2, 
                    text="Департамент по энергоэффективности Госстандарта", font=regular_font_10)
            
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=2, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=bottom_left)   
              
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=2, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=2, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=bottom_left) 
             
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=2, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+2, col_start=5, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')           
                                   
            set_cell(ws, row_start=start_row+3, col_start=5, col_end=7, 
                    text="(юридическое лицо)", merge_direction='horizontal', font = regular_font_9, alignment=center)      
                     
            set_cell(ws, row_start=start_row+4, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+5, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+7, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
                         
            set_cell(ws, row_start=start_row+9, col_start=5, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')               
                         
            set_cell(ws, row_start=start_row+10, col_start=5, col_end=7, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font = regular_font_9, row_height = 11)  
                                 
            set_cell(ws, row_start=start_row+11, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+12, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+13, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+14, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                 
        signatures_indicators_xlsx(ws, row_index + 1, plan)

        page_setttings(ws, print_area = "A1:G75")

        return ws

    def second_half_xlsx(wb, plan):
        ws = wb.create_sheet("Часть 2")
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 4), 
                ("B", 9), 
                ("C", 15), 
                ("D", 6),
                ("E", 6.75), 
                ("F", 6.43), 
                ("G", 6.86), 
                ("H", 6.43), 
                ("I", 9.43), 
                ("J", 5), 
                ("K", 7.86), 
                ("L", 7.43), 
                ("M", 6.57), 
                ("N", 6.57), 
                ("O", 6), 
                ("P", 7.14), 
                ("Q", 6.71), 
                ("R", 7)
                ]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21.75
            elif row == 4:
                ws.row_dimensions[row].height = 23.25
            elif row == 5:
                ws.row_dimensions[row].height = 19.5
            elif row == 6:
                ws.row_dimensions[row].height = 159
            else:
                ws.row_dimensions[row].height = 15

        ws.merge_cells("A1:R1")
        ws["A1"].value = "Часть 2. Мероприятия по экономии топливно-энергетических ресурсов"
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center

        ws.merge_cells("A2:R2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        ws.merge_cells("A3:A6"); ws["A3"].value = "№ п/п"
        ws.merge_cells("B3:B6"); ws["B3"].value = "Код основных направлений энергосбережения"
        ws.merge_cells("C3:C6"); ws["C3"].value = "Наименование мероприятия"
        ws.merge_cells("D3:D6"); ws["D3"].value = "Единицы измерения"
        ws.merge_cells("E3:E6"); ws["E3"].value = "Объем внедрения"
        ws.merge_cells("F3:G5"); ws["F3"].value = "Условно-годовой экономический эффект"
        ws["F6"].value = "т у.т."
        ws["G6"].value = "тыс. руб."
        ws.merge_cells("H3:H6"); ws["H3"].value = "Ожидаемый срок внедрения мероприятия, квартал"
        ws.merge_cells("I3:I6"); ws["I3"].value = "Ожидаемый экономический эффект от внедрения мероприятия в текущем году, т у.т."
        ws.merge_cells("J3:J6"); ws["J3"].value = "Срок окупаемости, лет"
        ws.merge_cells("K3:K6"); ws["K3"].value = "Объем финансирования, тыс. руб."
        ws.merge_cells("L3:R3"); ws["L3"].value = "источники финансирования, руб."
        ws.merge_cells("L4:O4"); ws["L4"].value = "бюджетные"
        ws.merge_cells("L5:L6"); ws["L5"].value = "республиканский бюджет на финансирование госпрограммы"
        ws.merge_cells("M5:M6"); ws["M5"].value = "республиканский бюджет"
        ws.merge_cells("N5:N6"); ws["N5"].value = "местный бюджет"
        ws.merge_cells("O5:O6"); ws["O5"].value = "другие"
        ws.merge_cells("P4:P6"); ws["P4"].value = "собственные средства организации"
        ws.merge_cells("Q4:Q6"); ws["Q4"].value = "кредиты банков, займы"
        ws.merge_cells("R4:R6"); ws["R4"].value = "иные"

        for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=18):
            for cell in row:
                if not cell.value:
                    continue
                if cell.coordinate in ("D3", "E3", "F6", "G6", "J3", "K3", "H3", "I3", "L5", "M5", "N5", "O5", "P4", "Q4", "R4"):
                    cell.alignment = vertical_text
                else:
                    cell.alignment = top
                cell.font = regular_font_10

        row_index = 7
        for col in range(1, 19):
            cell = ws.cell(row=row_index, column=col, value=col)
            cell.alignment = center
            cell.font = regular_font_10

        local_econ_execes = (EconExec.query
            .join(EconMeasure)
            .join(Plan)
            .filter(Plan.id == plan.id, EconExec.is_local == True)
            .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
            .all()
        )

        non_local_econ_execes = (EconExec.query
            .join(EconMeasure)
            .join(Plan)
            .filter(Plan.id == plan.id, EconExec.is_local == False)
            .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
            .all()
        )

        def add_section(title, execs, start_number=1):
            nonlocal row_index
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=18)
            ws.cell(row=row_index, column=1, value=title).font = bold_font_10
            ws.cell(row=row_index, column=1).alignment = center
            sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
            sums = {col: 0 for col in sum_cols}
            for idx, econ in enumerate(execs, start=start_number):
                row_index += 1
                row = [
                    idx,
                    econ.econ_measures.direction.code if econ.econ_measures and econ.econ_measures.direction else "",
                    econ.name if hasattr(econ, "name") else "",
                    econ.econ_measures.direction.unit.name if econ.econ_measures and econ.econ_measures.direction and econ.econ_measures.direction.unit else "",
                    econ.Volume if hasattr(econ, "Volume") else 0,
                    econ.EffTut if hasattr(econ, "EffTut") else 0,
                    econ.EffRub if hasattr(econ, "EffRub") else 0,
                    econ.ExpectedQuarter if hasattr(econ, "ExpectedQuarter") else 0,
                    econ.EffCurrYear if hasattr(econ, "EffCurrYear") else 0,
                    econ.Payback if hasattr(econ, "Payback") else 0,
                    econ.VolumeFin if hasattr(econ, "VolumeFin") else 0,
                    econ.BudgetState if hasattr(econ, "BudgetState") else 0,
                    econ.BudgetRep if hasattr(econ, "BudgetRep") else 0,
                    econ.BudgetLoc if hasattr(econ, "BudgetLoc") else 0,
                    econ.BudgetOther if hasattr(econ, "BudgetOther") else 0,
                    econ.MoneyOwn if hasattr(econ, "MoneyOwn") else 0,
                    econ.MoneyLoan if hasattr(econ, "MoneyLoan") else 0,
                    econ.MoneyOther if hasattr(econ, "MoneyOther") else 0
                ]
                for col in sum_cols:
                    try:
                        sums[col] += float(row[col-1])
                    except (TypeError, ValueError):
                        pass
                ws.append(row)
                for col_idx in range(1, 19):
                    cell = ws.cell(row=row_index, column=col_idx)
                    cell.alignment = left if col_idx == 3 else center
                    cell.font = regular_font_10
                    if col_idx in [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]:
                        cell.number_format = '0.000'
            row_index += 1
            
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value="Итого по разделу:").alignment = left
            ws.cell(row=row_index, column=1).font = regular_font_10_italic
            
            for col in sum_cols:
                cell = ws.cell(row=row_index, column=col, value=sums[col])
                cell.alignment = center
                cell.font = regular_font_10_italic
                cell.number_format = '0.000'
            return start_number + len(execs)

        next_number = add_section("Раздел 2. Мероприятия по экономии топливно-энергетических ресурсов", non_local_econ_execes, 1)
        # add_section("Раздел 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов", local_econ_execes, next_number)

        # row_index += 1
        
        # ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        # ws.cell(row=row_index, column=1, value="Всего по части 2:").font = bold_font_10
        # ws.cell(row=row_index, column=1).alignment = left
        
        # sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
        # total_sums = {col: 0 for col in sum_cols}
        # for row in ws.iter_rows(min_row=6, max_row=row_index, min_col=3, max_col=18):
        #     if row[0].value == "Итого по разделу:":
        #         for col in sum_cols:
        #             val = ws.cell(row=row[0].row, column=col).value
        #             try:
        #                 total_sums[col] += float(val)
        #             except (TypeError, ValueError):
        #                 pass
        # for col in sum_cols:
        #     cell = ws.cell(row=row_index, column=col, value=total_sums[col])
        #     cell.alignment = center
        #     cell.font = regular_font_10_italic
        #     cell.number_format = '0.000'

        quarters = [
            ("      январь–март", "jan_mar"),
            ("      январь–июнь", "jan_jun"),
            ("      январь–сентябрь", "jan_sep"),
            ("      январь–декабрь", "jan_dec")
        ]
        local_totals = get_cumulative_econ_metrics(plan.id, True)
        
        non_local_totals = get_cumulative_econ_metrics(plan.id, False)
        # total_metrics = {
        #     'jan_mar_eff': local_totals['jan_mar']['eff_curr_year'] + non_local_totals['jan_mar']['eff_curr_year'],
        #     'jan_mar_vol': local_totals['jan_mar']['volume_fin'] + non_local_totals['jan_mar']['volume_fin'],
        #     'jan_jun_eff': local_totals['jan_jun']['eff_curr_year'] + non_local_totals['jan_jun']['eff_curr_year'],
        #     'jan_jun_vol': local_totals['jan_jun']['volume_fin'] + non_local_totals['jan_jun']['volume_fin'],
        #     'jan_sep_eff': local_totals['jan_sep']['eff_curr_year'] + non_local_totals['jan_sep']['eff_curr_year'],
        #     'jan_sep_vol': local_totals['jan_sep']['volume_fin'] + non_local_totals['jan_sep']['volume_fin'],
        #     'jan_dec_eff': local_totals['jan_dec']['eff_curr_year'] + non_local_totals['jan_dec']['eff_curr_year'],
        #     'jan_dec_vol': local_totals['jan_dec']['volume_fin'] + non_local_totals['jan_dec']['volume_fin']
        # }
        
        total_metrics = {
            'jan_mar_eff': local_totals['jan_mar']['eff_curr_year'],
            'jan_mar_vol': local_totals['jan_mar']['volume_fin'],
            'jan_jun_eff': local_totals['jan_jun']['eff_curr_year'],
            'jan_jun_vol': local_totals['jan_jun']['volume_fin'],
            'jan_sep_eff': local_totals['jan_sep']['eff_curr_year'],
            'jan_sep_vol': local_totals['jan_sep']['volume_fin'],
            'jan_dec_eff': local_totals['jan_dec']['eff_curr_year'],
            'jan_dec_vol': local_totals['jan_dec']['volume_fin']
        }


        for q_label, q_key in quarters:
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value=q_label).font = regular_font_10
            ws.cell(row=row_index, column=1).alignment = left
            
            for col in range(3, 19):
                c = ws.cell(row=row_index, column=col)
                c.alignment = center
            
            eff_cell = ws.cell(row=row_index, column=9, value=total_metrics[f"{q_key}_eff"])
            eff_cell.number_format = '0.000'
            eff_cell.font = regular_font_10
            
            vol_cell = ws.cell(row=row_index, column=11, value=total_metrics[f"{q_key}_vol"])
            vol_cell.number_format = '0.000'
            vol_cell.font = regular_font_10 

        for row in ws.iter_rows(min_row=3, max_row=row_index, min_col=1, max_col=18):
            for cell in row:
                cell.border = thin_border

        def signatures_second_half_xlsx(ws, start_row, plan):
            big_space = " " * 10  
            text = f"Стоимость 1 т у.т. принята равной{big_space}долларов, при курсе{big_space}руб."

            set_cell(ws, row_start=start_row, col_start=2, col_end=9, 
                    text=text, font=regular_font_10_italic, row_height=None)

            set_cell(ws, row_start=start_row+2, col_start=2, row_end=start_row+3, col_end=9, text="3453")
      
      
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=3, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)   
              
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=3, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=3, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=center) 
             
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=3, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+2, col_start=12, col_end=17, 
                    text="__________________________________", merge_direction='horizontal')           
                                   
            set_cell(ws, row_start=start_row+3, col_start=12, col_end=16, 
                    text="(юридическое лицо)", merge_direction='horizontal', font = regular_font_9, alignment=center)      
                     
            set_cell(ws, row_start=start_row+4, col_start=12, col_end=15, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+5, col_start=12, col_end=15, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=12, col_end=15, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+7, col_start=12, col_end=15, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
                         
            set_cell(ws, row_start=start_row+9, col_start=12, col_end=17, 
                    text="__________________________________", merge_direction='horizontal')               
                         
            set_cell(ws, row_start=start_row+10, col_start=12, col_end=17, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font = regular_font_9, row_height = 11)  
                                 
            set_cell(ws, row_start=start_row+11, col_start=12, col_end=15, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+12, col_start=12, col_end=15, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+13, col_start=12, col_end=15, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+14, col_start=12, col_end=15, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
              
        signatures_second_half_xlsx(ws, row_index + 1, plan)
        page_setttings(ws, print_area="A1:R32")
        
        return ws

    def third_half_xlsx(wb, plan):
        ws = wb.create_sheet("Часть 3")
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 4), 
                ("B", 9), 
                ("C", 15), 
                ("D", 6),
                ("E", 6.75), 
                ("F", 6.43), 
                ("G", 6.86), 
                ("H", 6.43), 
                ("I", 9.43), 
                ("J", 5), 
                ("K", 7.86), 
                ("L", 7.43), 
                ("M", 6.57), 
                ("N", 6.57), 
                ("O", 6), 
                ("P", 7.14), 
                ("Q", 6.71), 
                ("R", 7)
                ]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21.75
            elif row == 4:
                ws.row_dimensions[row].height = 23.25
            elif row == 5:
                ws.row_dimensions[row].height = 26
            elif row == 6:
                ws.row_dimensions[row].height = 159
            else:
                ws.row_dimensions[row].height = 15

        ws.merge_cells("A1:R1")
        ws["A1"].value = "Часть 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов"
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center

        ws.merge_cells("A2:R2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        ws.merge_cells("A3:A6"); ws["A3"].value = "№ п/п"
        ws.merge_cells("B3:B6"); ws["B3"].value = "Код основных направлений энергосбережения"
        ws.merge_cells("C3:C6"); ws["C3"].value = "Наименование мероприятия"
        ws.merge_cells("D3:D6"); ws["D3"].value = "Единицы измерения"
        ws.merge_cells("E3:E6"); ws["E3"].value = "Объем внедрения"
        ws.merge_cells("F3:G5"); ws["F3"].value = "Условно-годовое увеличение использования местных ТЭР "
        ws["F6"].value = "т у.т."
        ws["G6"].value = "тыс. руб."
        ws.merge_cells("H3:H6"); ws["H3"].value = "Ожидаемый срок внедрения мероприятия, квартал"
        ws.merge_cells("I3:I6"); ws["I3"].value = "Ожидаемый экономический эффект от внедрения мероприятия в текущем году, т у.т."
        ws.merge_cells("J3:J6"); ws["J3"].value = "Срок окупаемости, лет"
        ws.merge_cells("K3:K6"); ws["K3"].value = "Объем финансирования, тыс. руб."
        ws.merge_cells("L3:R3"); ws["L3"].value = "источники финансирования, руб."
        ws.merge_cells("L4:O4"); ws["L4"].value = "бюджетные"
        ws.merge_cells("L5:L6"); ws["L5"].value = "республиканский бюджет на финансирование госпрограммы"
        ws.merge_cells("M5:M6"); ws["M5"].value = "республиканский бюджет"
        ws.merge_cells("N5:N6"); ws["N5"].value = "местный бюджет"
        ws.merge_cells("O5:O6"); ws["O5"].value = "другие"
        ws.merge_cells("P4:P6"); ws["P4"].value = "собственные средства организации"
        ws.merge_cells("Q4:Q6"); ws["Q4"].value = "кредиты банков, займы"
        ws.merge_cells("R4:R6"); ws["R4"].value = "иные"

        for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=18):
            for cell in row:
                if not cell.value:
                    continue
                if cell.coordinate in ("D3", "E3", "F6", "G6", "J3", "K3", "H3", "I3", "L5", "M5", "N5", "O5", "P4", "Q4", "R4"):
                    cell.alignment = vertical_text
                else:
                    cell.alignment = top
                cell.font = regular_font_10

        row_index = 7
        for col in range(1, 19):
            cell = ws.cell(row=row_index, column=col, value=col)
            cell.alignment = center
            cell.font = regular_font_10

        local_econ_execes = (EconExec.query
            .join(EconMeasure)
            .join(Plan)
            .filter(Plan.id == plan.id, EconExec.is_local == True)
            .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
            .all()
        )

        non_local_econ_execes = (EconExec.query
            .join(EconMeasure)
            .join(Plan)
            .filter(Plan.id == plan.id, EconExec.is_local == False)
            .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
            .all()
        )

        def add_section(title, execs, start_number=1):
            nonlocal row_index
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=18)
            ws.cell(row=row_index, column=1, value=title).font = bold_font_10
            ws.cell(row=row_index, column=1).alignment = center
            sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
            sums = {col: 0 for col in sum_cols}
            for idx, econ in enumerate(execs, start=start_number):
                row_index += 1
                row = [
                    idx,
                    econ.econ_measures.direction.code if econ.econ_measures and econ.econ_measures.direction else "",
                    econ.name if hasattr(econ, "name") else "",
                    econ.econ_measures.direction.unit.name if econ.econ_measures and econ.econ_measures.direction and econ.econ_measures.direction.unit else "",
                    econ.Volume if hasattr(econ, "Volume") else 0,
                    econ.EffTut if hasattr(econ, "EffTut") else 0,
                    econ.EffRub if hasattr(econ, "EffRub") else 0,
                    econ.ExpectedQuarter if hasattr(econ, "ExpectedQuarter") else 0,
                    econ.EffCurrYear if hasattr(econ, "EffCurrYear") else 0,
                    econ.Payback if hasattr(econ, "Payback") else 0,
                    econ.VolumeFin if hasattr(econ, "VolumeFin") else 0,
                    econ.BudgetState if hasattr(econ, "BudgetState") else 0,
                    econ.BudgetRep if hasattr(econ, "BudgetRep") else 0,
                    econ.BudgetLoc if hasattr(econ, "BudgetLoc") else 0,
                    econ.BudgetOther if hasattr(econ, "BudgetOther") else 0,
                    econ.MoneyOwn if hasattr(econ, "MoneyOwn") else 0,
                    econ.MoneyLoan if hasattr(econ, "MoneyLoan") else 0,
                    econ.MoneyOther if hasattr(econ, "MoneyOther") else 0
                ]
                for col in sum_cols:
                    try:
                        sums[col] += float(row[col-1])
                    except (TypeError, ValueError):
                        pass
                ws.append(row)
                for col_idx in range(1, 19):
                    cell = ws.cell(row=row_index, column=col_idx)
                    cell.alignment = left if col_idx == 3 else center
                    cell.font = regular_font_10
                    if col_idx in [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]:
                        cell.number_format = '0.000'
            row_index += 1
            
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value="Итого по разделу:").alignment = left
            ws.cell(row=row_index, column=1).font = regular_font_10_italic
            
            for col in sum_cols:
                cell = ws.cell(row=row_index, column=col, value=sums[col])
                cell.alignment = center
                cell.font = regular_font_10_italic
                cell.number_format = '0.000'
            return start_number + len(execs)

        add_section("Раздел 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов", local_econ_execes, 1)

        # row_index += 1
        
        # ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        # ws.cell(row=row_index, column=1, value="Всего по части 2:").font = bold_font_10
        # ws.cell(row=row_index, column=1).alignment = left
        
        # sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
        # total_sums = {col: 0 for col in sum_cols}
        # for row in ws.iter_rows(min_row=6, max_row=row_index, min_col=3, max_col=18):
        #     if row[0].value == "Итого по разделу:":
        #         for col in sum_cols:
        #             val = ws.cell(row=row[0].row, column=col).value
        #             try:
        #                 total_sums[col] += float(val)
        #             except (TypeError, ValueError):
        #                 pass
        # for col in sum_cols:
        #     cell = ws.cell(row=row_index, column=col, value=total_sums[col])
        #     cell.alignment = center
        #     cell.font = regular_font_10_italic
        #     cell.number_format = '0.000'

        quarters = [
            ("      январь–март", "jan_mar"),
            ("      январь–июнь", "jan_jun"),
            ("      январь–сентябрь", "jan_sep"),
            ("      январь–декабрь", "jan_dec")
        ]
        local_totals = get_cumulative_econ_metrics(plan.id, True)
        non_local_totals = get_cumulative_econ_metrics(plan.id, False)
        
        total_metrics = {
            'jan_mar_eff': non_local_totals['jan_mar']['eff_curr_year'],
            'jan_mar_vol': non_local_totals['jan_mar']['volume_fin'],
            'jan_jun_eff': non_local_totals['jan_jun']['eff_curr_year'],
            'jan_jun_vol': non_local_totals['jan_jun']['volume_fin'],
            'jan_sep_eff': non_local_totals['jan_sep']['eff_curr_year'],
            'jan_sep_vol': non_local_totals['jan_sep']['volume_fin'],
            'jan_dec_eff': non_local_totals['jan_dec']['eff_curr_year'],
            'jan_dec_vol': non_local_totals['jan_dec']['volume_fin']
        }

        for q_label, q_key in quarters:
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value=q_label).font = regular_font_10
            ws.cell(row=row_index, column=1).alignment = left
            
            for col in range(3, 19):
                c = ws.cell(row=row_index, column=col)
                c.alignment = center
            
            eff_cell = ws.cell(row=row_index, column=9, value=total_metrics[f"{q_key}_eff"])
            eff_cell.number_format = '0.000'
            eff_cell.font = regular_font_10
            
            vol_cell = ws.cell(row=row_index, column=11, value=total_metrics[f"{q_key}_vol"])
            vol_cell.number_format = '0.000'
            vol_cell.font = regular_font_10 

        for row in ws.iter_rows(min_row=3, max_row=row_index, min_col=1, max_col=18):
            for cell in row:
                cell.border = thin_border

        def signatures_third_half_xlsx(ws, start_row, plan):

            set_cell(ws, row_start=start_row+2, col_start=2, row_end=start_row+3, col_end=9, text='124')
    
    
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=3, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)   
            
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=3, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                    
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=3, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=center) 
            
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=3, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+2, col_start=12, col_end=17, 
                    text="__________________________________", merge_direction='horizontal')           
                                
            set_cell(ws, row_start=start_row+3, col_start=12, col_end=16, 
                    text="(юридическое лицо)", merge_direction='horizontal', font = regular_font_9, alignment=center)      
                    
            set_cell(ws, row_start=start_row+4, col_start=12, col_end=15, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                    
            set_cell(ws, row_start=start_row+5, col_start=12, col_end=15, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                    
            set_cell(ws, row_start=start_row+6, col_start=12, col_end=15, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                        
            set_cell(ws, row_start=start_row+7, col_start=12, col_end=15, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
                        
            set_cell(ws, row_start=start_row+9, col_start=12, col_end=17, 
                    text="__________________________________", merge_direction='horizontal')               
                        
            set_cell(ws, row_start=start_row+10, col_start=12, col_end=17, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font = regular_font_9, row_height = 11)  
                                
            set_cell(ws, row_start=start_row+11, col_start=12, col_end=15, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font = regular_font_9_italic, alignment=center)
                    
            set_cell(ws, row_start=start_row+12, col_start=12, col_end=15, 
                    text="_______________________", merge_direction='horizontal', row_height = 5, alignment=bottom_left)   
                    
            set_cell(ws, row_start=start_row+13, col_start=12, col_end=15, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font = regular_font_9, row_height = 11, alignment=left) 
                        
            set_cell(ws, row_start=start_row+14, col_start=12, col_end=15, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
            
        signatures_third_half_xlsx(ws, row_index + 1, plan)
        page_setttings(ws, print_area="A1:R32")
        
        return ws

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_title = title_xlsx(wb, plan)
    ws_firstHalf = first_half_xlsx(wb, plan)
    ws_secondHalf = second_half_xlsx(wb, plan)
    ws_thirdHalf = third_half_xlsx(wb, plan)
    
    wb.active = wb.index(ws_title)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"{plan.id}_{plan.year}.xlsx"

    return (
        file_stream,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename
    )
    